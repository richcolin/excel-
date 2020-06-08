#encoding:utf-8
import os
import glob
import openpyxl
import pandas as pd
def merge_xlsx_files(xlsx_files):
    #得到工作簿对象
    wb = openpyxl.Workbook()
    # wb=openpyxl.load_workbook(xlsx_files[0])
    ws = wb.active #获取活跃的工作表

    #汇总的工作表名称
    ws.title = "合并后的结果"
    num=int(input('正式数据从第几行开始:'))
    inpu_true = int(input('每行多于几个数据保留:'))
    #遍历所有的Excel文件
    for filename in xlsx_files:
        workbook = openpyxl.load_workbook(filename,data_only=True)
        #获取每个Excel文件中活跃的工作表
        sheet = workbook.active
        #按行获取所有单元格pip

        for row in sheet.iter_rows(min_row=num):
            values = [cell.value for cell in row]
            print(values)
            len_No=values.count(None)
            len_val=len(values)
            len_true=len_val-len_No
            print('len of No',values.count(None))
            print('len of val',len(values))
            # dis_values=set(values)
            # print(dis_values)
            # if dis_values=={None}:
            if len_true<inpu_true:
                break
            ws.append(values)
    return wb
def get_all_xlsx_files(path):
    xlsx_files = glob.glob(os.path.join(path, '*.xlsx'))

    sorted(xlsx_files, key=str.lower)
    return xlsx_files
def transformat():
    global path #定义为全局变量
    path = os.getcwd()#获取当前工作路径
    file = os.listdir(path)#获取当前路径下的所有文件
    for f in file:
        file_name_be,suff = os.path.splitext(f)#对路径进行分割，分别为文件路径和文件后缀
        if suff  == '.xls':
            print('将对{}文件进行转换...'.format(f))
            data = pd.DataFrame(pd.read_excel(path + '\\' + f))#读取xls文件
            data.to_excel(path + '\\' + file_name_be + '格式转变.xlsx',index = False)#格式转换
            print(' {} 文件已转化为 {} 保存在 {} 目录下\n'.format(f,file_name_be + '格式转变.xlsx',path))
transformat()
xlsx_files = get_all_xlsx_files(os.getcwd())
wb = merge_xlsx_files(xlsx_files)
#保存汇总后的工作簿文件
wb.save('合并后工作簿.xlsx')
